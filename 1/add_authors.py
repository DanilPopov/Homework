
from openpyxl import load_workbook
from dbs import db_session, User

def read_excel(filename):
    work_book = load_workbook(filename)
    work_sheet = work_book['Лист1']

    excel_data = list()
    for row in range(1, work_sheet.max_row):
        excel_row = dict()
        if work_sheet.cell(row=row, column=1).value is None:
            break
        excel_row['Headline'] = work_sheet.cell(row=row, column=1).value
        excel_row['URL'] = work_sheet.cell(row=row, column=2).value
        excel_row['Time'] = work_sheet.cell(row=row, column=3).value
        excel_row['Text'] = work_sheet.cell(row=row, column=4).value
        excel_row['email'] = work_sheet.cell(row=row, column=5).value
        excel_row['first_name'] = work_sheet.cell(row=row, column=6).value
        excel_row['last_name'] = work_sheet.cell(row=row, column=7).value
        excel_data.append(excel_row)

    return excel_data

# def show_excel_data(data):
#     print(data)

if __name__ == '__main__':
    excel_data = read_excel('Blog.xlsx')
    # show_excel_data(excel_data)

for row in excel_data:
    author = User(row['first_name'], row['last_name'], row['email'])
    db_session.add(author)

db_session.commit()